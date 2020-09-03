#智蜂考拉品类映射测试

import openpyxl
class ReadExcel(object):
    def __init__(self, excelPath):
        self.excelPath = excelPath
    def getData(self,sheetName = "Sheet2"):
        '''读取数据'''
        data = openpyxl.load_workbook(self.excelPath)
        table = data[sheetName]
        return table
        #获取第一行作为key值
        # self.keys = list(table.rows)[0]
        # self.rowNum = table.max_row  # 获取总行数
        # self.colNum = table.max_column  # 获取总列数
        # L = []
        # for i in range(5,self.rowNum):
        #     s = []
        #     for j in range(1,self.colNum + 1):
        #         value = table.cell(row=i,column=j).value
        #         s.append(value)
        #     L.append(s)
        # return L,len(L)
        # m_list = table.merged_cells
        # return m_listTypeError: int() argument must be a string, a bytes-like object or a number, not 'NoneType'


        # return value
        # if self.rowNum <= 1:
        #     print("总行数小于1")
        # else:
        #     L = []
        #     for i in range(2,self.rowNum + 1):
        #         s = {}
        #         for j in range(1,self.colNum + 1):
        #             rowValues = table.cell(row=i, column=j).value
        #             s[self.keys[j - 1].value] = rowValues
        #         L.append(s)
        #     return L
        # return list(table.rows)[0],table.max_row,table.max_column

if __name__ == '__main__':

    category = ReadExcel("s.xlsx")
    print(category.getData())
