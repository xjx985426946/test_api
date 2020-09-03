import openpyxl
class Excel(object):
    def __init__(self, excelPath):
        self.excelPath = excelPath
    def getData(self,sheetName = "Sheet1"):
        '''读取数据'''
        data = openpyxl.load_workbook(self.excelPath)
        table = data[sheetName]
        # 获取第一行作为key值
        self.keys = list(table.rows)[0]
        self.rowNum = table.max_row  # 获取总行数
        self.colNum = table.max_column  # 获取总列数
        if self.rowNum <= 1:
            print("总行数小于1")
        else:
            L = []
            for i in range(2,self.rowNum + 1):
                s = {}
                for j in range(1,self.colNum + 1):
                    rowValues = table.cell(row=i, column=j).value
                    s[self.keys[j - 1].value] = rowValues
                L.append(s)
            return L

    def copy_excel(self,excelpathNew):
        '''cope文件'''
        wb2 = openpyxl.Workbook()
        wb2.save(excelpathNew)
        # 读取数据
        wb1 = openpyxl.load_workbook(self.excelPath)
        wb2 = openpyxl.load_workbook(self.excelPath)
        sheets1 = wb1.sheetnames
        sheets2 = wb2.sheetnames
        sheet1 = wb1[sheets1[0]]
        sheet2 = wb2[sheets2[0]]
        max_row = sheet1.max_row
        max_column = sheet1.max_column
        for m in list(range(1, max_row + 1)):
            for n in list(range(97, 97 + max_column)):  # chr(97)='a'
                n = chr(n)  # ASCII字符
                i = '%s%d' % (n, m)  # 单元格编号
                cell1 = sheet1[i].value  # 获取data单元格数据
                sheet2[i].value = cell1  # 赋值到test单元格

        wb2.save(excelpathNew)  # 保存数据
        wb1.close()  # 关闭excel
        wb2.close()

    def writeData(self,sheet_name,row_n,col_n,value):
        '''修改e写入Excel数据'''
        wb =  openpyxl.load_workbook(self.excelPath)
        sheet = wb[sheet_name]
        sheet.cell(row=row_n, column=col_n).value = value
        wb.save(self.excelPath)
        wb.close()

    def readmore(self):
        wb1 = openpyxl.load_workbook(self.excelPath)
        sheets = wb1.sheetnames
        Ls = []
        for sheet in sheets:
            # print(sheet)
            table = wb1[sheet]
            # 获取第一行作为key值
            self.keys = list(table.rows)[0]
            self.rowNum = table.max_row  # 获取总行数
            self.colNum = table.max_column  # 获取总列数
            if self.rowNum <= 1:
                print("总行数小于1")
            else:
                L = []
                for i in range(2, self.rowNum + 1):
                    s = {}
                    for j in range(1, self.colNum + 1):
                        rowValues = table.cell(row=i, column=j).value
                        s[self.keys[j - 1].value] = rowValues
                    L.append(s)
            Ls.append(L)
        return Ls



